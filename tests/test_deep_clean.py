import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app.excel_service import ExcelService

def test_deep_clean_removes_named_ranges():
    # 1. Create a dummy Excel with named ranges and formatting
    test_file = "tests/temp_named_ranges.xlsx"
    if os.path.exists(test_file):
        os.remove(test_file)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = 10
    
    # Create a defined name (global)
    defn = openpyxl.workbook.defined_name.DefinedName("MyRange", attr_text="Sheet!$A$1")
    wb.defined_names.add(defn)
    
    # Add some formatting to verify it persists
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    wb.save(test_file)
    wb.close()
    
    # Verify initial state
    wb_check = openpyxl.load_workbook(test_file)
    print(f"DEBUG: Initial names: {list(wb_check.defined_names.keys())}")
    assert "MyRange" in wb_check.defined_names
    assert wb_check["Sheet"]["A1"].font.bold is True
    wb_check.close()
    
    # 2. Run Deep Clean
    service = ExcelService()
    success = service._deep_clean_excel(test_file)
    assert success is True
    
    # 3. Verify final state
    wb_final = openpyxl.load_workbook(test_file)
    print(f"DEBUG: Final names: {list(wb_final.defined_names.keys())}")
    assert "MyRange" not in wb_final.defined_names
    
    # Verify formatting still exists
    assert wb_final["Sheet"]["A1"].font.bold is True
    
    wb_final.close()
    
    # Cleanup
    if os.path.exists(test_file):
        os.remove(test_file)

if __name__ == "__main__":
    try:
        test_deep_clean_removes_named_ranges()
        print("Deep Clean Test Passed!")
    except Exception as e:
        print(f"Test Failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
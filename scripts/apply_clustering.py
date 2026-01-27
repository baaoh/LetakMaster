import sys
import os
import openpyxl

# Add project root to sys.path to import app
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.clustering import ProductClusterer

def apply_clustering(file_path, target_page):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    try:
        print(f"Loading {file_path}...")
        wb = openpyxl.load_workbook(file_path) # No data_only=True so we can write formulas if needed (though we write values)
        sheet = wb.active
        
        # 1. Read Data
        header_row = 6
        page_col = 21 # V
        name_col = 3  # D
        
        items = []
        rows_map = {} # Map item ID to Excel Row Index
        
        print(f"Scanning Page {target_page}...")
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
            if str(row[page_col]).strip() == str(target_page):
                name = row[name_col]
                if not name: continue
                
                item_id = row_idx
                items.append({
                    'id': item_id,
                    'name': name,
                    # We don't have weight/price in cols, extracting later
                })
                rows_map[item_id] = row_idx

        print(f"Found {len(items)} items. Clustering...")
        
        # 2. Cluster
        clusterer = ProductClusterer()
        groups = clusterer.group_items(items)
        
        print(f"Created {len(groups)} groups.")
        
        # 3. Write Back
        # Columns (0-indexed for internal logic, 1-indexed for openpyxl)
        # D (Name) -> 4
        # E (Doplnek) -> 5
        # G (EANY) -> 7
        # X (PSD_Allocation) -> 24
        
        COL_NAME = 4
        COL_DOPLNEK = 5
        COL_EANY = 7
        COL_ALLOC = 24
        
        for idx, grp in enumerate(groups, 1):
            leader = grp[0]
            leader_row = rows_map[leader['id']]
            
            # Generate Metadata
            names = [x['name'] for x in grp]
            nazev_a = clusterer.generate_smart_title(names)
            nazev_b = clusterer.generate_variants(grp, nazev_a)
            
            count = len(grp)
            label = "EAN: (check)" if count == 1 else (f"{count} druhy" if count <= 4 else "více druhů")
            
            group_id_str = f"A4_Grp_{idx:02d}"
            print(f"Writing Group {idx}: {group_id_str} | Title: {nazev_a} | Items: {count}")
            
            # Write Leader Data (Metadata + Allocation)
            sheet.cell(row=leader_row, column=COL_NAME).value = nazev_a
            sheet.cell(row=leader_row, column=COL_DOPLNEK).value = nazev_b
            sheet.cell(row=leader_row, column=COL_EANY).value = label
            sheet.cell(row=leader_row, column=COL_ALLOC).value = group_id_str
            
            # Write Children (Allocation ONLY)
            for child in grp[1:]:
                child_row = rows_map[child['id']]
                sheet.cell(row=child_row, column=COL_ALLOC).value = group_id_str
                # We do NOT overwrite name/doplnek for children, leaving them as original product data
        
        wb.save(file_path)
        print("Done. Saved changes.")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    apply_clustering(r"D:\TAMDA\LetakMaster\workspaces\state_1\Workspace_State_1.xlsx", 43)
